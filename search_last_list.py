from openpyxl import load_workbook

wb = load_workbook('temp.xlsx')


def last_page(col):

    wb.worksheets[0][f'{col}'] = None

    def find_edges(sheet):
        row = sheet.max_row
        while row > 0:
            cells = sheet[row]
            if all([cell.value is None for cell in cells]):
                row -= 1
            else:
                break
        if row == 0:
            return 0, 0

        column = sheet.max_column
        while column > 0:
            cells = next(sheet.iter_cols(min_col=column, max_col=column, max_row=row))
            if all([cell.value is None for cell in cells]):
                column -= 1
            else:
                break
        return row, column
    answ = find_edges(wb.worksheets[0])
    return answ[0]
