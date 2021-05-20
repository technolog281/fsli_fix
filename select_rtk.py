import pandas as pd
# from __init__ import *
# from encodings import ascii as us-ascii
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('xlsx/names_v5.xlsx')
sheet_1 = wb['Sheet1']
df = pd.DataFrame(
    columns=['G_NAME', 'NAME', 'CODE'])

for num in range(2, sheet_1.max_row + 1):
    values_2 = {'G_NAME': sheet_1[f'B{num}'].value,
                'NAME': sheet_1[f'C{num}'].value,
                'CODE': sheet_1[f'D{num}'].value}
    g_name = values_2['G_NAME']
    name = values_2['NAME']
    if 'производн' in g_name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'иммунолог' in name:
        values_2['CODE'] = '204'
        df = df.append(values_2, ignore_index=True)
    elif 'нтитела' in g_name:
        values_2['CODE'] = '204'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'Днк ' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'Рнк ' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'мплификац' in name:
        values_2['CODE'] = '207'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'икроскоп' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'иммуноферм' in name:
        values_2['CODE'] = '206'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'иммунохром' in name:
        values_2['CODE'] = '204'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'рН' in g_name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'онцентрация' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'одержание' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'ромбо' in g_name:
        values_2['CODE'] = '202'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'ейкоц' in g_name:
        values_2['CODE'] = '202'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'ритроц' in g_name:
        values_2['CODE'] = '202'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'имфоцит' in g_name:
        values_2['CODE'] = '202'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'иммунофлуор' in name:
        values_2['CODE'] = '207'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'иммунохеми' in name:
        values_2['CODE'] = '207'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'ремя' in name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'микроорган' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'ультур' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'язкост' in name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
    elif 'Запах' in name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
    elif 'Цвет' in name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
    elif 'олярная' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'скоро' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'бсолют' in name:
        values_2['CODE'] = '209'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'тносит' in name:
        values_2['CODE'] = '209'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'оксин' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
        continue
    elif 'дентификац' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
    elif 'бнаруж' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'тношени' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'рипп' in g_name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
    elif 'руппа кров' in g_name:
        values_2['CODE'] = '202'
        df = df.append(values_2, ignore_index=True)
    elif 'ph' in name:
        values_2['CODE'] = '200'
        df = df.append(values_2, ignore_index=True)
    elif 'итолог' in name:
        values_2['CODE'] = '209'
        df = df.append(values_2, ignore_index=True)
    elif 'минимальной подавл' in name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
    elif 'итолог' in g_name:
        values_2['CODE'] = '209'
        df = df.append(values_2, ignore_index=True)
    elif 'имницк' in g_name:
        values_2['CODE'] = '201'
        df = df.append(values_2, ignore_index=True)
    elif 'икробиолог' in g_name:
        values_2['CODE'] = '208'
        df = df.append(values_2, ignore_index=True)
    else:
        values_2['CODE'] = '000'
        df = df.append(values_2, ignore_index=True)

# print(rf)
df.to_excel('names_v9.xlsx')
